import openpyxl

def process_string(value):
    """Adds a space at the beginning of the string and removes leading zeros."""
    if value is None:
        return None
    # Add a space at the beginning and remove leading zeros
    return ' ' + str(value).lstrip('0')

def prepend_v_if_numeric(value):
    """Prepends 'v.' if the value contains only numbers."""
    if value is None:
        return None
    value_str = str(value).strip()
    if value_str.isdigit():
        return 'v.' + value_str
    return value

def replace_bk_with_book(value):
    """Replaces 'bk' with 'book'."""
    if value is None:
        return None
    if str(value).strip().lower() == 'bk':
        return 'book'
    return value

def append_period_to_abbreviations(value, abbreviations):
    """Appends a period to recognized abbreviations."""
    if value is None:
        return None
    value_str = str(value).strip().lower()
    if value_str in abbreviations:
        return value_str + '.'
    return value

def open_excel_file(file_path):
    abbreviations = {'pt', 'no'}
    
    try:
        # Load the Excel workbook
        wb = openpyxl.load_workbook(file_path)

        # Access the first worksheet (assuming it's the only one)
        sheet = wb.active

        # Perform the splitting and processing operation
        for row in sheet.iter_rows(min_row=1, max_col=12, values_only=False):
            # Process the 2nd column (index 1) to split its value by underscore
            if row[1].value is not None:
                split_values = str(row[1].value).split('_')

                # Write the first split value to the 10th column (index 9)
                if len(split_values) > 0:
                    row[9].value = split_values[0]

                # Write the second split value to the 11th column (index 10)
                if len(split_values) > 1:
                    value = replace_bk_with_book(split_values[1])
                    value = process_string(value)
                    value = prepend_v_if_numeric(value)
                    row[10].value = append_period_to_abbreviations(value, abbreviations)

                # Write the third split value to the 12th column (index 11)
                if len(split_values) > 2:
                    value = process_string(split_values[2])
                    row[11].value = value

        # Print each cell value in the row after processing
        for row in sheet.iter_rows(min_row=1, max_col=12, values_only=False):
            for cell in row:
                print(cell.value, end='\t')
            print()  # Print new line for each row

        # Save the changes to the workbook
        wb.save(file_path)

    except FileNotFoundError:
        print("File not found. Please provide a valid file path.")
    except openpyxl.utils.exceptions.InvalidFileException:
        print("Invalid file format. Please provide a valid Excel file.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Example usage
open_excel_file('Weblinks-public domain.xlsx')